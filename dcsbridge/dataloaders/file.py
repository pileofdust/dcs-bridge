import openpyxl
import re
import logging
from pathlib import Path
from . import DataLoader


class TextFileDataLoader(DataLoader):
    def __init__(self, data_file, headers=True, delimiter=";", encoding="UTF-8", columns: str = None):
        super().__init__()
        self.__data_file_path = Path(data_file)
        self.__headers = headers
        self.__delimiter = delimiter
        self.__encoding = encoding

        if columns:
            index, coords, alt = columns.split(",")
            self._INDEX_COLUMN = int(index) - 1
            self._COORDINATE_COLUMN = int(coords) - 1
            self._ALTITUDE_COLUMN = int(alt) - 1

    def load_data(self):
        self._reset_data()

        logging.info(f"Loading data from: {self.__data_file_path}")
        with open(self.__data_file_path, encoding=self.__encoding) as file:
            lines = file.readlines()

        if self.__headers and len(lines) > 1:
            lines = lines[1:]

        for line in lines:
            values = re.split(self.__delimiter, line.strip())
            self._add_row_as_waypoint(values)


class ExcelDataLoader(DataLoader):
    def __init__(self, data_file, headers=True):
        super().__init__()
        data_file_path = Path(data_file)
        wb = openpyxl.load_workbook(data_file_path)
        self.__ws = wb.worksheets[0]
        self.__headers = headers

    def load_data(self):
        self._reset_data()

        min_row = 2 if self.__headers else 1

        for row in self.__ws.iter_rows(min_row, self.__ws.max_row, 1, self.__ws.max_column, True):
            for cell in row:
                self._add_row_as_waypoint(row)
                break
