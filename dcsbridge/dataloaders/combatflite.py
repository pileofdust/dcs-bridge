import openpyxl
import re
from pathlib import Path


class MissionPlanDataLoader:
    ###
    #   'bullseye' = ("LAT", "LON"),
    #   'waypoint' = [(name, alt, lat, lon, notes)],
    #   'bingo' = "Bingo Value"
    ###

    # N36:59.650 E035:24.750
    __waypoint_coordinates_pattern = re.compile(r"([\w:.]+) ([\w:.]+)")

    # 25.0 A
    __altitude_pattern = re.compile(r"([\d.]+) \w")

    # N 35 00.000
    __bullseye_pattern = re.compile(r"([NE]) ([\d]+) ([\d.]+)")

    __replace_pattern = re.compile(r"[:. ]")

    def __reset_data(self):
        self.__bullseye = None
        self.__waypoints = {}
        self.__bingo = self.__default_bingo

    def __init__(self, path="data/Data.xlsx", bingo="2000"):
        data_file = Path(path)
        wb = openpyxl.load_workbook(data_file)
        self.__ws = wb.worksheets[0]

        self.__default_bingo = bingo
        self.__reset_data()

    def __bullseye(self, row):
        lat = row[0]
        lon = row[1]
        if lat and lon:
            coord = "%s%s%s" % self.__bullseye_pattern.match(lat).groups()
            self.__bullseye = self.__replace_pattern.sub("", coord)

    def __bingo(self, row):
        value = row[0]
        if value:
            self.__bingo = str(value)

    def __waypoint(self, row):
        altitude = row[1]

        if altitude:
            alt = self.__altitude_pattern.match(altitude).group(1)
        else:
            alt = "0"

        coordinates = row[4]
        lat, lon = self.__waypoint_coordinates_pattern.match(coordinates).groups()

        lat = self.__replace_pattern.sub("", lat)
        lon = self.__replace_pattern.sub("", lon)

        calculated_altitude = int(float(alt) * 1000)
        alt = str(calculated_altitude)
        self.__waypoints.append((lat, lon, alt))

    __commands = {"bullseye": __bullseye, "bingo": __bingo, "waypoint": __waypoint}

    def load_data(self):
        self.__reset_data()

        handler = None

        for row in self.__ws.iter_rows(1, self.__ws.max_row, 1, self.__ws.max_column, True):
            for cell in row:
                if not cell:
                    continue
                if str(cell).startswith("##"):
                    handler = self.__commands[cell[2:].lower()]
                elif handler:
                    handler(self, row)
                else:
                    print("Row not handled: ")
                    print(row)
                    print()

                break

    def get_waypoints(self):
        return self.__waypoints

    def get_bingo(self):
        return self.__bingo
