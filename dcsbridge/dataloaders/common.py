import openpyxl
import re
from pathlib import Path


class DataLoader:
    __INDEX_COLUMN = 0
    __COORDINATE_COLUMN = 1
    __ALTITUDE_COLUMN = 2

    ###
    # waypoint = (idx, lat / lon, alt)
    ###

    # 35°44.0705'N 37°06.23947'E
    __coordinates_pattern = re.compile(r"(\d+)[°˚](\d+).(\d+)\'([NS]) (\d+)[°˚](\d+).(\d+)\'([EW])")

    def __reset_data(self):
        self.__waypoints = {}

    def __init__(self, data_file):
        data_file_path = Path(data_file)
        wb = openpyxl.load_workbook(data_file_path)
        self.__ws = wb.worksheets[0]
        self.__waypoints = {}

    def __add_waypoint(self, row):
        idx = row[self.__INDEX_COLUMN]
        coordinates = row[self.__COORDINATE_COLUMN]
        altitude = row[self.__ALTITUDE_COLUMN]

        if not altitude:
            altitude = "0"
        else:
            altitude = round(altitude)

        t = self.__coordinates_pattern.match(coordinates).groups()
        lat = f"{t[3]}{t[0]}{t[1]}{t[2]}"
        lon = f"{t[7]}{int(t[4]):03d}{t[5]}{t[6]}"
        self.__waypoints[str(idx)] = (lat, lon, str(altitude))

    def load_data(self):
        self.__reset_data()

        handler = None

        for row in self.__ws.iter_rows(2, self.__ws.max_row, 1, self.__ws.max_column, True):
            for cell in row:
                self.__add_waypoint(row)
                break

    def get_waypoints(self):
        return self.__waypoints

    def get_waypoint(self, index):
        return self.__waypoints[index]
